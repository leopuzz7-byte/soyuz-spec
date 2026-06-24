"""
Генерация выходного Excel-файла спецификации.
"""

import io
import os
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from config import (
    COMPANY_HEADER, DIRECTOR_NAME, COMPANY_SHORT,
    SPEC_COLUMNS, PHOTO_ROW_HEIGHT_CM,
)

# Логотип, подпись, печать (необязательные файлы)
_LOGO_PATH      = Path(__file__).parent / 'logo.png'
_SIGNATURE_PATH = Path(__file__).parent / 'signature.png'
_STAMP_PATH     = Path(__file__).parent / 'stamp.png'

# ─── Стили ───────────────────────────────────────────────────────────────────
def _font(size=10, bold=False, color='000000', name='Arial'):
    return Font(name=name, size=size, bold=bold, color=color)

def _fill(color):
    return PatternFill(fill_type='solid', fgColor=color)

def _border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─── Главная функция ─────────────────────────────────────────────────────────
def build_excel(
    items: list[dict],
    object_name: str = '',
    spec_title: str = 'СПЕЦИФИКАЦИЯ по мебели',
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Спецификация'

    for col_name, col_letter, width in SPEC_COLUMNS:
        ws.column_dimensions[col_letter].width = width

    last_col = SPEC_COLUMNS[-1][1]   # 'J'
    current_row = 1

    # ── Строка 1: Шапка (текст слева, лого справа) ───────────────────────────
    ws.row_dimensions[current_row].height = 120

    if _LOGO_PATH.exists():
        # Текст: A:G, лого: H:J
        ws.merge_cells(f'A{current_row}:G{current_row}')
        tc = ws.cell(row=current_row, column=1)
        tc.value = COMPANY_HEADER
        tc.font = _font(size=8)
        tc.alignment = _align(h='center', v='center', wrap=True)
        tc.border = _border()

        ws.merge_cells(f'H{current_row}:{last_col}{current_row}')
        lc = ws.cell(row=current_row, column=8)
        lc.border = _border()
        lc.alignment = _align(h='center', v='center')

        try:
            from PIL import Image as PILImage
            with PILImage.open(str(_LOGO_PATH)) as pil_img:
                pil_img = pil_img.convert('RGBA')
                pil_img.thumbnail((280, 100), PILImage.LANCZOS)
                buf_logo = io.BytesIO()
                pil_img.save(buf_logo, 'PNG')
                buf_logo.seek(0)
            ws.add_image(XLImage(buf_logo), f'H{current_row}')
        except Exception as e:
            print(f'[excel] Логотип: {e}')
    else:
        # Без лого — весь ряд под текст
        ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
        tc = ws.cell(row=current_row, column=1)
        tc.value = COMPANY_HEADER
        tc.font = _font(size=8)
        tc.alignment = _align(h='center', v='center', wrap=True)
        tc.border = _border()

    current_row += 1

    # ── Строка 2: Заголовок спецификации ─────────────────────────────────────
    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = spec_title
    cell.font = _font(size=13, bold=True)
    cell.alignment = _align(h='center', v='center')
    cell.border = _border()
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # ── Строка 3: Объект ─────────────────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = f'Объект: {object_name}' if object_name else 'Объект:'
    cell.font = _font(size=10, bold=True)
    cell.alignment = _align(h='left', v='center')
    cell.border = _border()
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ── Строка 4: Заголовки колонок ──────────────────────────────────────────
    for col_idx, (col_name, _, _w) in enumerate(SPEC_COLUMNS, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = col_name
        cell.font = _font(size=9, bold=True)
        cell.fill = _fill('D9D9D9')   # светло-серый, не синий
        cell.alignment = _align(h='center', v='center', wrap=True)
        cell.border = _border()
    ws.row_dimensions[current_row].height = 42
    current_row += 1

    # ── Строки данных ─────────────────────────────────────────────────────────
    row_height_pt = PHOTO_ROW_HEIGHT_CM * 28.35
    total_sum = 0.0
    photo_positions = []

    for i, item in enumerate(items):
        retail_price = item.get('retail_price', 0) or 0
        qty = item.get('qty', 1) or 1
        total = retail_price * qty
        total_sum += total

        row_data = [
            item.get('row', i + 1),
            item.get('name', ''),
            item.get('description', ''),
            item.get('dimensions', ''),
            item.get('manufacturer', ''),
            item.get('article', ''),
            retail_price,
            qty,
            total,
            '',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.font = _font(size=9)
            cell.border = _border()
            if col_idx in (1, 7, 8, 9, 10):
                cell.alignment = _align(h='center', v='center')
            else:
                cell.alignment = _align(h='left', v='center')
            if col_idx in (7, 9):
                cell.number_format = '#,##0.00'

        ws.row_dimensions[current_row].height = row_height_pt

        if item.get('photo_bytes'):
            photo_positions.append((current_row, item['photo_bytes'], 10))

        current_row += 1

    # ── Строка ИТОГО ─────────────────────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell_label = ws.cell(row=current_row, column=1)
    cell_label.value = 'ИТОГО К ОПЛАТЕ С НДС:'
    cell_label.font = _font(size=10, bold=True)
    cell_label.alignment = _align(h='right', v='center')
    cell_label.border = _border()

    cell_total = ws.cell(row=current_row, column=9)
    cell_total.value = total_sum
    cell_total.font = _font(size=10, bold=True)
    cell_total.alignment = _align(h='center', v='center')
    cell_total.border = _border()
    cell_total.number_format = '#,##0.00'

    ws.cell(row=current_row, column=10).border = _border()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # ── Подпись ───────────────────────────────────────────────────────────────
    sig_row = current_row
    row_h = 50 if _SIGNATURE_PATH.exists() else 24

    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws.cell(row=current_row, column=1).value = f'Генеральный директор {COMPANY_SHORT}'
    ws.cell(row=current_row, column=1).font = _font(size=9)
    ws.cell(row=current_row, column=1).alignment = _align(h='left', v='center', wrap=False)

    ws.merge_cells(f'D{current_row}:G{current_row}')
    if not _SIGNATURE_PATH.exists():
        ws.cell(row=current_row, column=4).value = '____________________________'
        ws.cell(row=current_row, column=4).font = _font(size=9)
        ws.cell(row=current_row, column=4).alignment = _align(h='center', v='bottom', wrap=False)

    ws.merge_cells(f'H{current_row}:{last_col}{current_row}')
    ws.cell(row=current_row, column=8).value = DIRECTOR_NAME
    ws.cell(row=current_row, column=8).font = _font(size=9)
    ws.cell(row=current_row, column=8).alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[current_row].height = row_h
    current_row += 1

    # Строка: (подпись) / (расшифровка)
    ws.merge_cells(f'D{current_row}:G{current_row}')
    ws.cell(row=current_row, column=4).value = '(подпись)'
    ws.cell(row=current_row, column=4).font = _font(size=7, color='888888')
    ws.cell(row=current_row, column=4).alignment = _align(h='center', v='top', wrap=False)

    ws.merge_cells(f'H{current_row}:{last_col}{current_row}')
    ws.cell(row=current_row, column=8).value = '(расшифровка подписи)'
    ws.cell(row=current_row, column=8).font = _font(size=7, color='888888')
    ws.cell(row=current_row, column=8).alignment = _align(h='left', v='top', wrap=False)
    ws.row_dimensions[current_row].height = 14
    current_row += 1

    # Строка: М.П. / печать
    stamp_row = current_row
    ws.merge_cells(f'A{current_row}:C{current_row}')
    if _STAMP_PATH.exists():
        ws.row_dimensions[current_row].height = 80
    else:
        ws.cell(row=current_row, column=1).value = 'М.П.'
        ws.cell(row=current_row, column=1).font = _font(size=9)
        ws.cell(row=current_row, column=1).alignment = _align(h='left', v='center', wrap=False)
        ws.row_dimensions[current_row].height = 20

    # ── Вставка фотографий товаров ────────────────────────────────────────────
    tmp_files = _insert_photos(ws, photo_positions, row_height_pt)

    # ── Вставка подписи ───────────────────────────────────────────────────────
    if _SIGNATURE_PATH.exists():
        try:
            sig_img = XLImage(str(_SIGNATURE_PATH))
            sig_w = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(4, 8)) * 7.5
            sig_h = row_h * 1.33
            scale = min(sig_w / (sig_img.width or 200), sig_h / (sig_img.height or 60), 1.0)
            sig_img.width  = max(int((sig_img.width or 200) * scale), 40)
            sig_img.height = max(int((sig_img.height or 60) * scale), 20)
            ws.add_image(sig_img, f'D{sig_row}')
        except Exception as e:
            print(f'[excel] Подпись: {e}')

    # ── Вставка печати ────────────────────────────────────────────────────────
    if _STAMP_PATH.exists():
        try:
            stamp_img = XLImage(str(_STAMP_PATH))
            st_w = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(1, 4)) * 7.5
            st_h = 80 * 1.33
            scale = min(st_w / (stamp_img.width or 200), st_h / (stamp_img.height or 200), 1.0)
            stamp_img.width  = max(int((stamp_img.width or 200) * scale), 40)
            stamp_img.height = max(int((stamp_img.height or 200) * scale), 40)
            ws.add_image(stamp_img, f'A{stamp_row}')
        except Exception as e:
            print(f'[excel] Печать: {e}')

    # ── Параметры печати (без freeze panes) ──────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)

    for p in tmp_files:
        try:
            os.unlink(p)
        except Exception:
            pass

    return buf.getvalue()


def _insert_photos(ws, photo_positions: list, row_height_pt: float) -> list:
    if not photo_positions:
        return []

    tmp_files = []
    for row_idx, photo_bytes, col_idx in photo_positions:
        try:
            fd, tmp_path = tempfile.mkstemp(suffix='.png')
            with os.fdopen(fd, 'wb') as f:
                f.write(photo_bytes)
            tmp_files.append(tmp_path)

            img = XLImage(tmp_path)
            cell_w = ws.column_dimensions[get_column_letter(col_idx)].width * 7.5
            cell_h = row_height_pt * 1.33
            orig_w = img.width  or 100
            orig_h = img.height or 100
            scale = min((cell_w - 8) / orig_w, (cell_h - 8) / orig_h, 1.0)
            img.width  = max(int(orig_w * scale), 20)
            img.height = max(int(orig_h * scale), 20)
            ws.add_image(img, f'{get_column_letter(col_idx)}{row_idx}')
        except Exception as e:
            print(f'[excel] Фото строка {row_idx}: {e}')

    return tmp_files
