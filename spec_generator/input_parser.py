"""
Парсинг входного Excel-файла с списком мебели.

Поддерживает разные форматы входных файлов:
- Обязательные: строка с «Наименование» (или похожее) + столбцы данных
- Необязательные: №, Ед.изм., Количество, Цена, Сумма
"""

import io
import openpyxl


# Синонимы для поиска колонок (нижний регистр)
COL_SYNONYMS = {
    'name':  ['наименование', 'наименовние', 'название', 'товар', 'позиция', 'описание'],
    'qty':   ['количество', 'кол-во', 'кол.', 'кол', 'шт', 'штук'],
    'price': ['цена', 'стоимость', 'цена,руб', 'цена руб'],
    'unit':  ['ед.изм', 'ед. изм', 'единица', 'ед'],
    'num':   ['№', '№п/п', 'n', 'номер', 'п/п', '#'],
}


def _find_header_row(ws) -> tuple[int, dict]:
    """
    Ищет строку заголовков в листе.
    Возвращает (row_idx, col_map) где col_map = {'name': col_idx, ...}
    row_idx — 1-based номер строки заголовка.
    """
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_values = [
            str(ws.cell(row=row_idx, column=c).value or '').strip().lower()
            for c in range(1, ws.max_column + 1)
        ]
        col_map = {}
        for field, synonyms in COL_SYNONYMS.items():
            for c_idx, val in enumerate(row_values, start=1):
                if any(syn in val for syn in synonyms):
                    col_map[field] = c_idx
                    break
        if 'name' in col_map:
            return row_idx, col_map
    return -1, {}


def parse_input_excel(file_bytes: bytes) -> tuple[str, list[dict]]:
    """
    Парсит входной Excel.
    Возвращает (title, items) где:
      title — заголовок спецификации (первая непустая строка до заголовков)
      items — список {'row': int, 'name': str, 'qty': int,
                       'budget_price': float, 'unit': str}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row, col_map = _find_header_row(ws)
    if header_row == -1 or 'name' not in col_map:
        raise ValueError(
            'Не найдена колонка «Наименование» во входном файле. '
            'Убедитесь, что файл содержит таблицу с колонкой «Наименование».'
        )

    # Заголовок документа — первая непустая строка до строки заголовков
    title = ''
    for row_idx in range(1, header_row):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col).value
            if val and str(val).strip():
                title = str(val).strip()
                break
        if title:
            break

    # Читаем данные
    items = []
    row_counter = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=col_map['name'])
        name = str(name_cell.value or '').strip()

        if not name or name.upper() in ('ИТОГО', 'ИТОГО:', 'ВСЕГО', 'ИТОГО К ОПЛАТЕ'):
            continue

        # Количество
        qty = 1
        if 'qty' in col_map:
            qty_val = ws.cell(row=row_idx, column=col_map['qty']).value
            try:
                qty = int(float(str(qty_val))) if qty_val is not None else 1
            except (ValueError, TypeError):
                qty = 1

        # Цена (бюджетная, из входного файла)
        budget_price = 0.0
        if 'price' in col_map:
            price_val = ws.cell(row=row_idx, column=col_map['price']).value
            try:
                budget_price = float(price_val) if price_val is not None else 0.0
            except (ValueError, TypeError):
                budget_price = 0.0

        # Единица измерения
        unit = 'шт'
        if 'unit' in col_map:
            unit_val = ws.cell(row=row_idx, column=col_map['unit']).value
            if unit_val:
                unit = str(unit_val).strip()

        row_counter += 1
        items.append({
            'row':          row_counter,
            'name':         name,
            'qty':          qty,
            'unit':         unit,
            'budget_price': budget_price,
        })

    if not items:
        raise ValueError('В файле не найдено ни одной позиции.')

    return title, items
